from typing import List, Optional, Dict, Any, Tuple
import os
import uuid
import openpyxl
from io import BytesIO
from datetime import datetime, timezone
from fastapi import HTTPException, UploadFile
from pymongo import InsertOne, UpdateOne
from ..repositories.upload_repository import UploadRepository
from ..core.config import get_settings
from ..core.security import hash_password
from ..utils.auth_utils import generate_student_password
from ..core.audit import log_audit

class UploadService:
    def __init__(self):
        self.repo = UploadRepository()
        self.settings = get_settings()

    async def upload_student_document(
        self, 
        student_id: str, 
        document_type: str, 
        file: UploadFile, 
        user: Dict[str, Any]
    ) -> Dict[str, Any]:
        student = await self.repo.get_student_by_id(student_id)
        if not student:
            raise HTTPException(status_code=404, detail="Student not found")
        
        allowed_types = ["application/pdf", "image/jpeg", "image/png", "image/jpg"]
        if file.content_type not in allowed_types:
            raise HTTPException(status_code=400, detail="Invalid file type. Only PDF and images allowed.")
        
        contents = await file.read()
        if len(contents) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File too large. Max 5MB allowed.")
        
        ext = file.filename.split(".")[-1] if "." in file.filename else "pdf"
        filename = f"{student_id}_{document_type}_{uuid.uuid4().hex[:8]}.{ext}"
        
        folder_map = {
            "tenth_certificate": "certificates",
            "twelfth_certificate": "certificates",
            "id_proof": "id_proofs",
            "photo": "documents"
        }
        folder = folder_map.get(document_type, "documents")
        
        upload_path = self.settings.uploads_dir / folder
        upload_path.mkdir(parents=True, exist_ok=True)
        file_path = upload_path / filename
        with open(file_path, "wb") as f:
            f.write(contents)
        
        doc_record = {
            "id": str(uuid.uuid4()),
            "student_id": student_id,
            "document_type": document_type,
            "file_name": file.filename,
            "file_path": f"/api/upload/files/{folder}/{filename}",
            "file_size": len(contents),
            "mime_type": file.content_type,
            "uploaded_by": user["id"],
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "is_verified": False
        }
        await self.repo.insert_student_document(doc_record)
        await self.repo.update_student_profile_doc(student_id, document_type, doc_record["file_path"])
        
        await log_audit(user["id"], "upload", document_type, doc_record["id"], after_value={"student_id": student_id, "file_name": file.filename}, user_name=user["name"], user_role=user["role"], module="documents")
        
        return doc_record
    
    async def upload_issued_document(
        self, 
        registration_number: str, 
        document_type: str, 
        file: UploadFile, 
        user: Dict[str, Any]
    ) -> str:
        contents = await file.read()
        ext = file.filename.split(".")[-1] if "." in file.filename else "pdf"
        filename = f"issued_{registration_number}_{document_type}_{uuid.uuid4().hex[:8]}.{ext}"
        
        folder = "issued_documents"
        upload_path = self.settings.uploads_dir / folder
        upload_path.mkdir(parents=True, exist_ok=True)
        file_path = upload_path / filename
        
        with open(file_path, "wb") as f:
            f.write(contents)
            
        return f"/api/upload/files/{folder}/{filename}"

    async def upload_students_excel(self, file: UploadFile, user: Dict[str, Any]) -> Dict[str, Any]:
        if not file.filename.endswith('.xlsx'):
            raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")
        
        contents = await file.read()
        try:
            wb = openpyxl.load_workbook(BytesIO(contents))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")
            
        ws = wb.active
        headers = [str(cell.value).strip().lower() for cell in ws[1] if cell.value]
        required_cols = ['email', 'name', 'password', 'roll_number', 'department_code', 'batch']
        
        for req in required_cols:
            if req not in headers:
                raise HTTPException(status_code=400, detail=f"Missing required column: {req}")
        
        col_map = {h: i for i, h in enumerate(headers)}
        dept_map = await self.repo.get_departments_map()
        
        total_rows = 0
        success_count = 0
        failed_count = 0
        errors = []
        user_ops = []
        student_ops = []
        
        def normalize_quota(val: str) -> str:
            if not val: return val
            v = str(val).lower().strip()
            if "7.5" in v: return "7.5 Quota"
            if "pmss" in v: return "PMSS"
            if "fg" in v or "first graduate" in v: return "FG"
            if "government" in v: return "Government Quota"
            if "management" in v: return "Management Quota"
            if v == "mq": return "Management Quota"
            return val.strip()

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            total_rows += 1
            try:
                row_data = {h: row[col_map[h]] for h in headers if col_map[h] < len(row)}
                email = str(row_data.get('email', '')).strip()
                name = str(row_data.get('name', '')).strip()
                password = str(row_data.get('password', '')).strip()
                roll_number = str(row_data.get('roll_number', '')).strip()
                dept_code = str(row_data.get('department_code', '')).strip()
                batch = str(row_data.get('batch', '')).strip()
                
                if not all([email, name, password, roll_number, dept_code, batch]):
                    failed_count += 1
                    errors.append({"row": row_num, "error": "Missing required data in fields"})
                    continue
                    
                dept_id = dept_map.get(dept_code)
                if not dept_id:
                    failed_count += 1
                    errors.append({"row": row_num, "error": f"Department code not found: {dept_code}"})
                    continue
                
                user_id = str(uuid.uuid4())
                user_doc = {
                    "id": user_id,
                    "email": email,
                    "name": name,
                    "password": await hash_password(generate_student_password(name)),
                    "role": "student",
                    "department_id": dept_id,
                    "is_active": True,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "permissions": [],
                    "sub_roles": []
                }
                
                student_id = str(uuid.uuid4())
                student_doc = {
                    "id": student_id,
                    "user_id": user_id,
                    "roll_number": roll_number,
                    "department_id": dept_id,
                    "batch": batch,
                    "semester": int(row_data.get('semester', 1) or 1),
                    "section": row_data.get('section'),
                    "regulation": row_data.get('regulation', 'R2023'),
                    "admission_date": datetime.now(timezone.utc).isoformat(),
                    "is_active": True,
                    "cgpa": 0.0,
                    # Extended mapping
                    "admission_quota": normalize_quota(row_data.get('admission_quota')),
                    "community": row_data.get('community'),
                    "admission_type": row_data.get('admission_type'),
                    "parent_name": row_data.get('parent_name'),
                    "parent_phone": str(row_data.get('parent_mobile', row_data.get('parent_phone', ''))).strip(),
                    "blood_group": row_data.get('blood_group'),
                    "permanent_address": row_data.get('address', row_data.get('permanent_address')),
                    "gender": row_data.get('gender'),
                    "date_of_birth": str(row_data.get('date_of_birth', '')) if row_data.get('date_of_birth') else None,
                    "mobile_number": str(row_data.get('mobile_number', '')).strip()
                }
                user_ops.append(InsertOne(user_doc))
                student_ops.append(InsertOne(student_doc))
                success_count += 1
            except Exception as e:
                failed_count += 1
                errors.append({"row": row_num, "error": str(e)})
        
        if user_ops:
            await self.repo.bulk_write_users_students(user_ops, student_ops)

        await self.repo.create_notification({
            "id": str(uuid.uuid4()),
            "user_id": user["id"],
            "role": user["role"],
            "type": "system",
            "title": "Student Upload Complete",
            "message": f"Successfully added {success_count} students. {failed_count} failed.",
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        return {
            "message": f"Bulk student upload completed. {success_count} success, {failed_count} failed.",
            "total_processed": total_rows,
            "success": success_count,
            "failed": failed_count,
            "errors": errors
        }

    async def upload_marks_excel(self, file: UploadFile, user: Dict[str, Any]) -> Dict[str, Any]:
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Only Excel files are allowed")
        
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        required = ['roll_number', 'subject_code', 'academic_year', 'semester', 'exam_type', 'marks']
        
        for req in required:
            if req not in headers:
                raise HTTPException(status_code=400, detail=f"Missing required column: {req}")
        
        col_map = {h: i for i, h in enumerate(headers)}
        success_count = 0
        errors = []
        marks_ops = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                data = {h: row[col_map[h]] for h in headers if col_map[h] < len(row)}
                if not data.get('roll_number'): continue
                
                student = await self.repo.get_student_by_roll(str(data['roll_number']))
                if not student:
                    errors.append({"row": row_num, "error": f"Student not found: {data['roll_number']}"})
                    continue
                
                subject = await self.repo.get_subject_by_code(str(data['subject_code']))
                if not subject:
                    errors.append({"row": row_num, "error": f"Subject not found: {data['subject_code']}"})
                    continue
                
                academic_year = str(data['academic_year'])
                semester = int(data['semester'])
                exam_type = str(data['exam_type']).lower()
                marks_val = float(data['marks'])
                
                existing = await self.repo.get_marks_record(student["id"], subject["id"], academic_year, semester)
                if existing:
                    if existing.get("is_locked"):
                        errors.append({"row": row_num, "error": "Marks locked"})
                        continue
                    marks_ops.append(UpdateOne(
                        {"id": existing["id"]},
                        {"$set": {
                            exam_type: marks_val,
                            "updated_by": user["id"],
                            "updated_at": datetime.now(timezone.utc).isoformat()
                        }}
                    ))
                else:
                    marks_ops.append(InsertOne({
                        "id": str(uuid.uuid4()),
                        "student_id": student["id"],
                        "subject_id": subject["id"],
                        "academic_year": academic_year,
                        "semester": semester,
                        exam_type: marks_val,
                        "is_locked": False,
                        "updated_by": user["id"],
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }))
                success_count += 1
            except Exception as e:
                errors.append({"row": row_num, "error": str(e)})
                
        if marks_ops:
            await self.repo.bulk_write_marks(marks_ops)
        return {"message": f"Bulk marks upload completed: {success_count} success", "errors": errors}

    async def upload_attendance_excel(self, file: UploadFile, user: Dict[str, Any]) -> Dict[str, Any]:
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Only Excel files are allowed")
        
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        required = ['roll_number', 'subject_code', 'date', 'status']
        
        for req in required:
            if req not in headers:
                raise HTTPException(status_code=400, detail=f"Missing required column: {req}")
        
        col_map = {h: i for i, h in enumerate(headers)}
        success_count = 0
        errors = []
        attendance_ops = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                data = {h: row[col_map[h]] for h in headers if col_map[h] < len(row)}
                if not data.get('roll_number'): continue
                
                student = await self.repo.get_student_by_roll(str(data['roll_number']))
                if not student:
                    errors.append({"row": row_num, "error": f"Student not found: {data['roll_number']}"})
                    continue
                
                subject = await self.repo.get_subject_by_code(str(data['subject_code']))
                if not subject:
                    errors.append({"row": row_num, "error": f"Subject not found: {data['subject_code']}"})
                    continue
                
                date_val = data['date']
                date_str = date_val.strftime('%Y-%m-%d') if isinstance(date_val, datetime) else str(date_val)
                
                attendance_doc = {
                    "id": str(uuid.uuid4()),
                    "student_id": student["id"],
                    "subject_id": subject["id"],
                    "date": date_str,
                    "status": str(data['status']).lower(),
                    "marked_by": user["id"],
                    "marked_at": datetime.now(timezone.utc).isoformat()
                }
                attendance_ops.append(UpdateOne(
                    {"student_id": student["id"], "subject_id": subject["id"], "date": date_str},
                    {"$set": attendance_doc},
                    upsert=True
                ))
                success_count += 1
            except Exception as e:
                errors.append({"row": row_num, "error": str(e)})
                
        if attendance_ops:
            await self.repo.bulk_write_attendance(attendance_ops)
        return {"message": f"Bulk attendance upload completed: {success_count} success", "errors": errors}
